import openpyxl
import json
import os

filepath = r"C:\Aby\magang\PLTA Mrica\Bank Sampah\Project bank sampah PLN\data yang diperlukan\data excel\26 PS Rekapitulasi Program Pengelolaan Sampah PLTA Wonogiri.xlsx"

wb = openpyxl.load_workbook(filepath, data_only=True)

print(f"=== FILE: {os.path.basename(filepath)} ===")
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheet names: {wb.sheetnames}")
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: '{sheet_name}'")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}")
    print(f"{'='*80}")
    
    # Print all rows with data (up to 200 rows to be thorough)
    max_rows_to_print = min(ws.max_row, 200)
    for row_idx in range(1, max_rows_to_print + 1):
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is not None:
                row_data.append(f"[C{col_idx}] {val}")
        if row_data:
            print(f"  Row {row_idx}: {' | '.join(row_data)}")
    
    if ws.max_row > 200:
        print(f"  ... (showing first 200 of {ws.max_row} rows)")

print("\n\n=== ANALYSIS COMPLETE ===")
